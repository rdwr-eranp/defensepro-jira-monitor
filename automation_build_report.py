"""
Automation Build Report - DefensePro
Generates an HTML report of automation test results for a version from a
specified minimum build number.

Usage:
    python automation_build_report.py [--version 10.13.0.0] [--min-build 90]
"""

import os, sys, html, argparse
from datetime import datetime
import psycopg2
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

VERSION     = '10.13.0.0'
MIN_BUILD   = 90
JIRA_URL    = 'https://rwrnd.atlassian.net'

PLATFORM_TYPE_MAP = {
    'UHT':   'FPGA',     'MRQP':  'FPGA',     'MR2':   'FPGA',
    'ESXI':  'Software', 'KVM':   'Software', 'VL3':   'Software', 'HT2': 'Software',
    'MRQ_X': 'EZchip',  'MRQX':  'EZchip',
}

ALL_PT_MODES = [f'{pt} - {m}' for pt in ['EZchip', 'FPGA', 'Software']
                               for m in ['Routing', 'Transparent']]

KNOWN_PLATFORMS = ('UHT','MRQP','MR2','ESXI','KVM','VL3','HT2','MRQ_X','MRQX')


def get_prior_versions(version):
    """Return baseline versions for coverage comparison."""
    parts = [int(part) for part in version.split('.')]
    major, minor = parts[0], parts[1]
    tail = parts[2:] or [0, 0]
    zero_tail = '.'.join('0' for _ in tail)

    if any(tail):
        return [f'{major}.{minor}.{zero_tail}']

    return [f'{major}.{v}.{zero_tail}' for v in range(minor - 1, max(minor - 3, -1), -1)]


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def connect():
    return psycopg2.connect(
        host=os.getenv('PG_HOST', '10.185.20.124'),
        port=os.getenv('PG_PORT', '5432'),
        database=os.getenv('PG_DATABASE', 'results'),
        user=os.getenv('PG_USER', 'postgres'),
        password=os.getenv('PG_PASSWORD', '')
    )


def query(conn, sql):
    return pd.read_sql(sql, conn)


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def get_build_summary(conn, version, min_build):
    """Per-build: executions, passed, failed, pass_ratio"""
    sql = f"""
        SELECT
            te.build,
            COUNT(1)                                                         AS executions,
            SUM(CASE WHEN LOWER(te.status) = 'passed'                  THEN 1 ELSE 0 END) AS passed,
            SUM(CASE WHEN LOWER(te.status) IN ('failed','error','fail') THEN 1 ELSE 0 END) AS failed,
            MIN(te.start_time)::date AS first_run,
            MAX(te.start_time)::date AS last_run
        FROM test_execution te
        WHERE te.version = '{version}'
          AND te.build >= {min_build}
          AND te.mode  = 'regression'
        GROUP BY te.build
        ORDER BY te.build
    """
    df = query(conn, sql)
    df['pass_ratio'] = df['passed'] / df['executions'].clip(lower=1) * 100
    return df


def get_latest_status_by_platform_type(conn, version, min_build):
    """Latest result per (test_id, individual platform, mode) across all builds >= min_build."""
    sql = f"""
        WITH ranked AS (
            SELECT
                te.test_id,
                t.name  AS test_name,
                d.platform,
                CASE
                    WHEN d.platform IN ('UHT','MRQP','MR2')        THEN 'FPGA'
                    WHEN d.platform IN ('ESXI','KVM','VL3','HT2')  THEN 'Software'
                    WHEN d.platform IN ('MRQ_X','MRQX')            THEN 'EZchip'
                    ELSE 'Other'
                END AS platform_type,
                CASE WHEN p.name LIKE '%-Routing' THEN 'Routing' ELSE 'Transparent' END AS mode,
                LOWER(te.status) AS status,
                te.build,
                ROW_NUMBER() OVER (
                    PARTITION BY te.test_id, d.platform,
                        CASE WHEN p.name LIKE '%-Routing' THEN 'Routing' ELSE 'Transparent' END
                    ORDER BY te.start_time DESC
                ) AS rn
            FROM test_execution te
            JOIN device  d ON te.device_id  = d.id
            JOIN profile p ON te.profile_id = p.id
            JOIN test    t ON te.test_id    = t.id
            WHERE te.version = '{version}'
              AND te.build  >= {min_build}
              AND te.mode    = 'regression'
        )
        SELECT test_id, test_name, platform, platform_type, mode, status, build
        FROM ranked
        WHERE rn = 1
          AND platform_type != 'Other'
    """
    df = query(conn, sql)
    df['pt_mode'] = df['platform_type'] + ' - ' + df['mode']
    return df


def get_available_tests_baseline(conn, prior_versions):
    """Count distinct tests per platform_type+mode from prior releases (baseline)."""
    ver_list = "', '".join(prior_versions)
    sql = f"""
        SELECT
            CASE
                WHEN d.platform IN ('UHT','MRQP','MR2')       THEN 'FPGA'
                WHEN d.platform IN ('ESXI','KVM','VL3','HT2') THEN 'Software'
                WHEN d.platform IN ('MRQ_X','MRQX')           THEN 'EZchip'
                ELSE 'Other'
            END AS platform_type,
            CASE WHEN p.name LIKE '%-Routing' THEN 'Routing' ELSE 'Transparent' END AS mode,
            COUNT(DISTINCT te.test_id) AS available_tests
        FROM test_execution te
        JOIN device  d ON te.device_id  = d.id
        JOIN profile p ON te.profile_id = p.id
        WHERE te.version IN ('{ver_list}')
          AND te.mode = 'regression'
        GROUP BY 1, 2
    """
    df = query(conn, sql)
    df['pt_mode'] = df['platform_type'] + ' - ' + df['mode']
    return df


def get_new_tests(conn, version, min_build, prior_versions):
    """Find tests in the current version that did not appear in any prior version."""
    ver_list = "', '".join(prior_versions)
    plat_list = "', '".join(KNOWN_PLATFORMS)
    sql = f"""
        WITH current_tests AS (
            SELECT DISTINCT
                te.test_id,
                t.name AS test_name,
                CASE
                    WHEN d.platform IN ('UHT','MRQP','MR2')       THEN 'FPGA'
                    WHEN d.platform IN ('ESXI','KVM','VL3','HT2') THEN 'Software'
                    WHEN d.platform IN ('MRQ_X','MRQX')           THEN 'EZchip'
                END AS platform_type,
                CASE WHEN p.name LIKE '%-Routing' THEN 'Routing' ELSE 'Transparent' END AS mode
            FROM test_execution te
            JOIN device  d ON te.device_id  = d.id
            JOIN profile p ON te.profile_id = p.id
            JOIN test    t ON te.test_id    = t.id
            WHERE te.version = '{version}'
              AND te.build  >= {min_build}
              AND te.mode    = 'regression'
              AND d.platform IN ('{plat_list}')
        )
        SELECT ct.test_id, ct.test_name, ct.platform_type, ct.mode
        FROM current_tests ct
        WHERE ct.test_id NOT IN (
            SELECT DISTINCT test_id
            FROM test_execution
            WHERE version IN ('{ver_list}')
              AND mode = 'regression'
        )
        ORDER BY ct.platform_type, ct.mode, ct.test_name
    """
    df = query(conn, sql)
    if not df.empty:
        df['pt_mode'] = df['platform_type'] + ' - ' + df['mode']
    return df


def get_critical_failures(conn, version, min_build):
    """Tests failing on more than one platform in the latest runs."""
    sql = f"""
        WITH ranked AS (
            SELECT
                te.test_id,
                t.name AS test_name,
                d.platform,
                LOWER(te.status) AS status,
                ROW_NUMBER() OVER (
                    PARTITION BY te.test_id, d.platform
                    ORDER BY te.start_time DESC
                ) AS rn
            FROM test_execution te
            JOIN device  d ON te.device_id  = d.id
            JOIN profile p ON te.profile_id = p.id
            JOIN test    t ON te.test_id    = t.id
            WHERE te.version = '{version}'
              AND te.build  >= {min_build}
              AND te.mode    = 'regression'
              AND d.platform NOT IN ('Other')
        ),
        latest AS (
            SELECT test_id, test_name, platform, status
            FROM ranked WHERE rn = 1
        ),
        agg AS (
            SELECT test_id, test_name,
                   COUNT(DISTINCT platform)                                              AS total_platforms,
                   SUM(CASE WHEN status IN ('failed','error','fail') THEN 1 ELSE 0 END) AS failed_platforms
            FROM latest
            GROUP BY test_id, test_name
        )
        SELECT test_id, test_name, total_platforms, failed_platforms
        FROM agg
        WHERE failed_platforms > 1
        ORDER BY failed_platforms DESC, test_name
    """
    return query(conn, sql)


# ---------------------------------------------------------------------------
# HTML generation helpers
# ---------------------------------------------------------------------------

def pct_bar(value, color='#4caf50', width=120):
    pct = min(max(value, 0), 100)
    return (
        f'<div style="display:inline-block;width:{width}px;background:#e0e0e0;'
        f'border-radius:4px;overflow:hidden;vertical-align:middle;">'
        f'<div style="width:{pct:.0f}%;height:14px;background:{color};"></div></div>'
        f'&nbsp;<span style="font-size:11px;">{value:.1f}%</span>'
    )


def pass_color(ratio):
    if ratio >= 90: return '#4caf50'
    if ratio >= 70: return '#ff9800'
    return '#dc3545'


def build_html(version, min_build, build_df, latest_df, baseline_df, critical_df, new_tests_df, prior_versions):
    now = datetime.now().strftime('%d %b %Y, %H:%M')
    builds_list = ', '.join(str(b) for b in sorted(build_df['build'].tolist()))
    total_execs  = int(build_df['executions'].sum())   # raw total across all builds (CI activity)
    # Headline pass/fail counts use last result per (test_id, platform, mode) — no double-counting
    dedup_total  = len(latest_df)
    total_passed = int((latest_df['status'] == 'passed').sum())
    total_failed = int(latest_df['status'].isin(['failed', 'error', 'fail']).sum())
    overall_pass = total_passed / max(dedup_total, 1) * 100
    n_builds     = len(build_df)

    # --- Platform Type table ---
    # Deduplicate: one result per (test_id, platform_type, mode) — pick the highest-build run.
    # This prevents double-counting tests that ran on multiple devices of the same type
    # (e.g., a test on both ESXI and KVM both count as "Software").
    pt_latest_df = (
        latest_df
        .sort_values('build', ascending=False)
        .drop_duplicates(subset=['test_id', 'platform_type', 'mode'])
    )

    pt_rows = ''
    for pt_mode in ALL_PT_MODES:
        sub = pt_latest_df[pt_latest_df['pt_mode'] == pt_mode]
        baseline_row = baseline_df[baseline_df['pt_mode'] == pt_mode]
        available    = int(baseline_row['available_tests'].sum()) if not baseline_row.empty else 0
        unique_tests = sub['test_id'].nunique()
        passed       = int((sub['status'] == 'passed').sum())
        failed       = int(sub['status'].isin(['failed', 'error', 'fail']).sum())
        ratio        = passed / max(unique_tests, 1) * 100
        coverage     = unique_tests / max(available, 1) * 100

        color = pass_color(ratio)
        pt_rows += f"""
        <tr>
            <td><strong>{html.escape(pt_mode)}</strong></td>
            <td style="text-align:center;">{unique_tests}</td>
            <td style="text-align:center;">{available if available else '-'}</td>
            <td>{pct_bar(coverage, '#6c5ce7', 100) if available else '-'}</td>
            <td style="text-align:center;color:#4caf50;font-weight:bold;">{passed}</td>
            <td style="text-align:center;color:#dc3545;font-weight:bold;">{failed}</td>
            <td>{pct_bar(ratio, color, 100)}</td>
        </tr>"""

    # --- Per-Platform table (individual devices) ---
    # Determine all (platform, mode) combos present, sorted by platform_type then platform then mode
    ORDER_PT = {'EZchip': 0, 'FPGA': 1, 'Software': 2, 'Other': 3}
    combos_df = (latest_df[['platform', 'platform_type', 'mode']]
                 .drop_duplicates()
                 .copy())
    combos_df['_pt_ord']  = combos_df['platform_type'].map(ORDER_PT).fillna(9)
    combos_df['_mode_ord'] = combos_df['mode'].map({'Routing': 0, 'Transparent': 1}).fillna(2)
    combos_df = combos_df.sort_values(['_pt_ord', 'platform', '_mode_ord'])

    plat_rows = ''
    for _, combo in combos_df.iterrows():
        plat     = combo['platform']
        pt       = combo['platform_type']
        mode     = combo['mode']
        pt_mode  = f'{pt} - {mode}'
        sub      = latest_df[(latest_df['platform'] == plat) & (latest_df['mode'] == mode)]
        bline    = baseline_df[baseline_df['pt_mode'] == pt_mode]
        available     = int(bline['available_tests'].sum()) if not bline.empty else 0
        unique_tests  = sub['test_id'].nunique()
        passed        = int((sub['status'] == 'passed').sum())
        failed        = int(sub['status'].isin(['failed', 'error', 'fail']).sum())
        ratio         = passed / max(unique_tests, 1) * 100
        coverage      = unique_tests / max(available, 1) * 100
        color         = pass_color(ratio)
        plat_rows += f"""
        <tr>
            <td><strong>{html.escape(plat)}</strong></td>
            <td style="text-align:center;color:#888;font-size:12px;">{html.escape(pt)} / {html.escape(mode)}</td>
            <td style="text-align:center;">{unique_tests}</td>
            <td style="text-align:center;">{available if available else '-'}</td>
            <td>{pct_bar(coverage, '#6c5ce7', 100) if available else '-'}</td>
            <td style="text-align:center;color:#4caf50;font-weight:bold;">{passed}</td>
            <td style="text-align:center;color:#dc3545;font-weight:bold;">{failed}</td>
            <td>{pct_bar(ratio, color, 100)}</td>
        </tr>"""

    # --- Per-build table ---
    build_rows = ''
    for _, row in build_df.iterrows():
        ratio = row['pass_ratio']
        color = pass_color(ratio)
        build_rows += f"""
        <tr>
            <td style="text-align:center;font-weight:bold;">{int(row['build'])}</td>
            <td style="text-align:center;">{row['first_run']}</td>
            <td style="text-align:center;">{row['last_run']}</td>
            <td style="text-align:center;">{int(row['executions'])}</td>
            <td style="text-align:center;color:#4caf50;font-weight:bold;">{int(row['passed'])}</td>
            <td style="text-align:center;color:#dc3545;font-weight:bold;">{int(row['failed'])}</td>
            <td>{pct_bar(ratio, color, 120)}</td>
        </tr>"""

    # --- Critical failures ---
    critical_rows = ''
    if not critical_df.empty:
        for _, row in critical_df.iterrows():
            critical_rows += (
                f'<tr>'
                f'<td>{html.escape(str(row["test_id"]))}</td>'
                f'<td>{html.escape(str(row["test_name"]))}</td>'
                f'<td style="text-align:center;color:#dc3545;font-weight:bold;">{int(row["failed_platforms"])}</td>'
                f'<td style="text-align:center;">{int(row["total_platforms"])}</td>'
                f'</tr>'
            )
    else:
        critical_rows = '<tr><td colspan="4" style="text-align:center;color:#4caf50;">\u2705 No tests failing on more than one platform</td></tr>'

    # --- Plotly build trend (inline) ---
    builds_js  = str(sorted(build_df['build'].tolist()))
    passed_js  = str([int(build_df[build_df['build'] == b]['passed'].values[0]) for b in sorted(build_df['build'].tolist())])
    failed_js  = str([int(build_df[build_df['build'] == b]['failed'].values[0]) for b in sorted(build_df['build'].tolist())])
    ratio_js   = str([round(build_df[build_df['build'] == b]['pass_ratio'].values[0], 1) for b in sorted(build_df['build'].tolist())])

    # --- New tests section ---
    prior_ver_str = ', '.join(prior_versions)
    n_new_tests = new_tests_df['test_id'].nunique() if not new_tests_df.empty else 0
    if not new_tests_df.empty:
        new_summary_rows = ''
        for pt_mode in ALL_PT_MODES:
            sub = new_tests_df[new_tests_df['pt_mode'] == pt_mode]
            count = sub['test_id'].nunique()
            if count:
                new_summary_rows += (
                    f'<tr>'
                    f'<td><strong>{html.escape(pt_mode)}</strong></td>'
                    f'<td style="text-align:center;">{count}</td>'
                    f'</tr>'
                )
        new_detail_rows = ''
        for _, row in new_tests_df.drop_duplicates('test_id').sort_values('test_name').iterrows():
            new_detail_rows += (
                f'<tr>'
                f'<td>{html.escape(str(row["test_id"]))}</td>'
                f'<td>{html.escape(str(row["test_name"]))}</td>'
                f'</tr>'
            )
        new_tests_html = (
            f'<h2>&#x1F195; New Tests (Not in Prior Releases: {html.escape(prior_ver_str)})</h2>'
            f'<p>These <strong>{n_new_tests}</strong> tests are new to this release and not yet included in '
            f'the coverage baseline. They will be part of the baseline for the next release.</p>'
            f'<table><thead><tr><th>Platform Type / Mode</th><th>New Test Count</th></tr></thead>'
            f'<tbody>{new_summary_rows}</tbody></table>'
            f'<details><summary style="cursor:pointer;color:#4472C4;font-weight:bold;margin:12px 0;">'
            f'&#9654; Show all {n_new_tests} new test details</summary>'
            f'<table style="margin-top:8px;"><thead><tr>'
            f'<th style="width:130px;">Test ID</th><th>Test Name</th>'
            f'</tr></thead><tbody>{new_detail_rows}</tbody></table></details>'
        )
    else:
        new_tests_html = (
            f'<h2>&#x1F195; New Tests (Not in Prior Releases: {html.escape(prior_ver_str)})</h2>'
            f'<p style="color:#4caf50;">&#10003; No new tests &#8212; all executed tests were present in prior releases.</p>'
        )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Automation Report – {version} Build ≥{min_build}</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  body      {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; margin: 20px 40px; }}
  h1        {{ color: #2c3e50; border-bottom: 3px solid #4472C4; padding-bottom: 8px; }}
  h2        {{ color: #4472C4; margin-top: 36px; }}
  h3        {{ color: #2c3e50; margin-top: 24px; }}
  table     {{ border-collapse: collapse; width: 100%; margin-bottom: 16px; }}
  th        {{ background: #4472C4; color: white; padding: 8px 12px; text-align: left; }}
  td        {{ border: 1px solid #ddd; padding: 7px 10px; vertical-align: middle; }}
  tr:nth-child(even) {{ background: #f7f7f7; }}
  .kpi-row  {{ display: flex; gap: 20px; flex-wrap: wrap; margin: 20px 0; }}
  .kpi      {{ flex: 1; min-width: 160px; padding: 16px 20px; border-radius: 8px;
               color: white; text-align: center; }}
  .kpi-num  {{ font-size: 32px; font-weight: bold; }}
  .kpi-lbl  {{ font-size: 13px; margin-top: 4px; opacity: .9; }}
  .banner   {{ background: linear-gradient(135deg,#2c3e50,#4472C4); color:white;
               padding: 16px 24px; border-radius: 8px; margin-bottom: 24px; }}
</style>
</head>
<body>

<div class="banner">
  <h1 style="color:white;border:none;margin:0;">🤖 Automation Status Report</h1>
  <p style="margin:4px 0 0;">DefensePro <strong>{version}</strong> &nbsp;|&nbsp;
     Builds <strong>≥ {min_build}</strong> ({builds_list}) &nbsp;|&nbsp; Generated {now}</p>
</div>

<div class="kpi-row">
  <div class="kpi" style="background:linear-gradient(135deg,#00b894,#00cec9);">
    <div class="kpi-num">{n_builds}</div><div class="kpi-lbl">Builds Included</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,#6c5ce7,#a29bfe);"> 
    <div class="kpi-num">{total_execs:,}</div><div class="kpi-lbl">Total Executions (Raw)</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,#4caf50,#8bc34a);">
    <div class="kpi-num">{total_passed:,}</div><div class="kpi-lbl">Passed (Last Result)</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,#dc3545,#e17055);">
    <div class="kpi-num">{total_failed:,}</div><div class="kpi-lbl">Failed (Last Result)</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,{'#4caf50,#8bc34a' if overall_pass >= 90 else '#ff9800,#fdcb6e' if overall_pass >= 70 else '#dc3545,#e17055'});">
    <div class="kpi-num">{overall_pass:.1f}%</div><div class="kpi-lbl">Overall Pass Rate</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,#e74c3c,#c0392b);">
    <div class="kpi-num">{len(critical_df)}</div><div class="kpi-lbl">Critical Failures</div>
  </div>
  <div class="kpi" style="background:linear-gradient(135deg,#0984e3,#74b9ff);">
    <div class="kpi-num">{n_new_tests}</div><div class="kpi-lbl">New Tests Added</div>
  </div>
</div>

<h2>📈 Pass Ratio Trend by Build</h2>
<div id="buildChart" style="height:360px;"></div>
<script>
var builds  = {builds_js};
var passed  = {passed_js};
var failed  = {failed_js};
var ratios  = {ratio_js};
Plotly.newPlot('buildChart', [
  {{type:'bar', name:'Passed', x: builds, y: passed, marker:{{color:'#4caf50'}}}},
  {{type:'bar', name:'Failed', x: builds, y: failed, marker:{{color:'#dc3545'}}}},
  {{type:'scatter', name:'Pass %', x: builds, y: ratios, yaxis:'y2',
    mode:'lines+markers', line:{{color:'#6c5ce7', width:2}},
    marker:{{size:6}}}},
], {{
  barmode:'stack',
  yaxis:  {{title:'Executions', gridcolor:'#eee'}},
  yaxis2: {{title:'Pass %', overlaying:'y', side:'right', range:[0,105], ticksuffix:'%'}},
  legend: {{orientation:'h', y:-0.2}},
  margin: {{l:60,r:60,t:20,b:60}},
  plot_bgcolor:'#fafafa', paper_bgcolor:'white'
}});
</script>

<h2>🖥 Results by Platform Type (Unique Tests, Latest Result per Test)</h2>
<table>
  <thead>
    <tr>
      <th>Platform Type / Mode</th>
      <th>Unique Tests</th>
      <th>Baseline Available</th>
      <th>Coverage</th>
      <th style="color:#a8f0b0;">Passed</th>
      <th style="color:#ffb3b3;">Failed</th>
      <th>Pass Ratio</th>
    </tr>
  </thead>
  <tbody>{pt_rows}</tbody>
</table>

<h2>� Results by Individual Platform (Latest per Test)</h2>
<table>
  <thead>
    <tr>
      <th>Platform</th>
      <th>Type / Mode</th>
      <th>Unique Tests</th>
      <th>Baseline</th>
      <th>Coverage</th>
      <th style="color:#a8f0b0;">Passed</th>
      <th style="color:#ffb3b3;">Failed</th>
      <th>Pass Ratio</th>
    </tr>
  </thead>
  <tbody>{plat_rows}</tbody>
</table>

<h2>�📋 Per-Build Summary</h2>
<table>
  <thead>
    <tr>
      <th>Build</th><th>First Run</th><th>Last Run</th>
      <th>Executions</th><th style="color:#a8f0b0;">Passed</th>
      <th style="color:#ffb3b3;">Failed</th><th>Pass Ratio</th>
    </tr>
  </thead>
  <tbody>{build_rows}</tbody>
</table>

<h2>🚨 Critical Failures (Failing on More Than One Platform)</h2>
<table>
  <thead><tr><th style="width:120px;">Test ID</th><th>Test Name</th><th style="width:130px;">Failed Platforms</th><th style="width:130px;">Total Platforms</th></tr></thead>
  <tbody>{critical_rows}</tbody>
</table>

{new_tests_html}

<p style="color:#999;font-size:12px;margin-top:40px;">
  Report generated {now} &nbsp;|&nbsp; version {version} &nbsp;|&nbsp;
  builds ≥ {min_build} &nbsp;|&nbsp; mode: regression
</p>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def export_critical_failures_excel(critical_df, latest_df, build_df, baseline_df, version, min_build):
    """Export critical failures to Excel with per-platform status breakdown."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    xlsx_file = f'critical_failures_{version.replace(".", "_")}_build{min_build}plus_{ts}.xlsx'

    # --- Build detail grid: one row per test, one col per (platform, mode) ---
    crit_ids = set(critical_df['test_id'].tolist())
    detail = latest_df[latest_df['test_id'].isin(crit_ids)].copy()

    # Unique (platform, mode) combos, sorted
    combos = sorted(detail[['platform', 'mode']].drop_duplicates()
                    .apply(lambda r: f"{r['platform']} / {r['mode']}", axis=1).tolist())

    # Pivot: index = (test_id, test_name), columns = "platform / mode", values = status
    # Also bring in failed_platforms count
    detail['col'] = detail['platform'] + ' / ' + detail['mode']
    pivot = (detail.pivot_table(index=['test_id', 'test_name'], columns='col',
                                values='status', aggfunc='first')
                   .reset_index())
    pivot.columns.name = None

    # Merge failed/total platform counts
    crit_counts = critical_df[['test_id', 'failed_platforms', 'total_platforms']]
    pivot = pivot.merge(crit_counts, on='test_id', how='left')
    # Move counts right after test_name
    cols = ['test_id', 'test_name', 'failed_platforms', 'total_platforms'] + \
           [c for c in pivot.columns if c not in ('test_id','test_name','failed_platforms','total_platforms')]
    pivot = pivot[cols]

    # --- Build per-build summary df ---
    build_summary = build_df[['build', 'executions', 'passed', 'failed',
                               'pass_ratio', 'first_run', 'last_run']].copy()
    build_summary['pass_ratio'] = build_summary['pass_ratio'].round(1)

    # --- Colors ---
    RED    = PatternFill('solid', fgColor='FFCCCC')
    GREEN  = PatternFill('solid', fgColor='CCFFCC')
    ORANGE = PatternFill('solid', fgColor='FFE0B2')
    BLUE   = PatternFill('solid', fgColor='D0E8FF')
    GRAY   = PatternFill('solid', fgColor='F2F2F2')
    HDR    = PatternFill('solid', fgColor='4472C4')

    hdr_font   = Font(bold=True, color='FFFFFF')
    bold_font  = Font(bold=True)
    thin_side  = Side(style='thin', color='BBBBBB')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    def status_fill(s):
        s = str(s).lower() if s else ''
        if s == 'passed':           return GREEN
        if s in ('failed','error','fail'): return RED
        if s == 'skipped':          return ORANGE
        return GRAY

    wb = openpyxl.Workbook()

    # ── Sheet 1: Critical Failures Detail ──────────────────────────────────
    ws = wb.active
    ws.title = 'Critical Failures'

    header = ['Test ID', 'Test Name', 'Failed Platforms', 'Total Platforms'] + sorted(combos)
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # pivot columns: test_id, test_name, failed_platforms, total_platforms, ...combos...
    STATUS_COL_START = 5  # first status column (1-based)
    for ri, row in enumerate(pivot.itertuples(index=False), 2):
        vals = list(row)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=(ci == 2))
            if ci == 1:
                cell.font = Font(bold=True, color='1155CC',
                                 underline='single')
                # Make test_id a hyperlink
                cell.hyperlink = f'https://rwrnd.atlassian.net/browse/{val}'
            elif ci in (3, 4):  # failed_platforms / total_platforms
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if ci == 3:  # failed_platforms — highlight in red
                    cell.font = Font(bold=True, color='CC0000')
            elif ci >= STATUS_COL_START:
                cell.fill = status_fill(val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Zebra (fixed columns: test_id, test_name, failed_platforms, total_platforms)
        if ri % 2 == 0:
            for ci in range(1, 5):
                if ws.cell(ri, ci).fill.patternType is None or \
                   ws.cell(ri, ci).fill.fgColor.rgb == '00000000':
                    ws.cell(ri, ci).fill = GRAY

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 16  # Failed Platforms
    ws.column_dimensions['D'].width = 16  # Total Platforms
    for ci in range(5, len(header) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Per Platform ──────────────────────────────────────────────
    ORDER_PT = {'EZchip': 0, 'FPGA': 1, 'Software': 2, 'Other': 3}
    combos_df = (latest_df[['platform', 'platform_type', 'mode']]
                 .drop_duplicates().copy())
    combos_df['_pt_ord']   = combos_df['platform_type'].map(ORDER_PT).fillna(9)
    combos_df['_mode_ord'] = combos_df['mode'].map({'Routing': 0, 'Transparent': 1}).fillna(2)
    combos_df = combos_df.sort_values(['_pt_ord', 'platform', '_mode_ord'])

    ws_pp = wb.create_sheet('Per Platform', 1)  # insert as 2nd sheet
    hdr_pp = ['Platform', 'Type', 'Mode', 'Unique Tests', 'Baseline',
               'Coverage %', 'Passed', 'Failed', 'Pass Ratio %']
    for ci, h in enumerate(hdr_pp, 1):
        cell = ws_pp.cell(row=1, column=ci, value=h)
        cell.fill = HDR; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, (_, combo) in enumerate(combos_df.iterrows(), 2):
        plat, pt, mode = combo['platform'], combo['platform_type'], combo['mode']
        pt_mode = f'{pt} - {mode}'
        sub   = latest_df[(latest_df['platform'] == plat) & (latest_df['mode'] == mode)]
        bline = baseline_df[baseline_df['pt_mode'] == pt_mode]
        available    = int(bline['available_tests'].sum()) if not bline.empty else 0
        unique_tests = sub['test_id'].nunique()
        passed_cnt   = int((sub['status'] == 'passed').sum())
        failed_cnt   = int(sub['status'].isin(['failed', 'error', 'fail']).sum())
        ratio_val    = round(passed_cnt / max(unique_tests, 1) * 100, 1)
        cov_val      = round(unique_tests / max(available, 1) * 100, 1) if available else None
        pp_vals = [plat, pt, mode, unique_tests,
                   available if available else 'N/A',
                   cov_val if cov_val is not None else 'N/A',
                   passed_cnt, failed_cnt, ratio_val]
        for ci, val in enumerate(pp_vals, 1):
            cell = ws_pp.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if ci != 1 else 'left', vertical='center')
            if ci == 9:  # Pass Ratio
                r = float(val)
                cell.fill = GREEN if r >= 90 else ORANGE if r >= 70 else RED
                cell.font = Font(bold=True)
            elif ci == 7:  # Passed
                cell.fill = PatternFill('solid', fgColor='CCFFCC')
            elif ci == 8:  # Failed
                cell.fill = PatternFill('solid', fgColor='FFCCCC')
        if ri % 2 == 0:
            for ci in range(1, 4):
                if ws_pp.cell(ri, ci).fill.fgColor.rgb == '00000000':
                    ws_pp.cell(ri, ci).fill = GRAY

    for ci, w in enumerate([12, 12, 14, 14, 12, 12, 10, 10, 14], 1):
        ws_pp.column_dimensions[get_column_letter(ci)].width = w
    ws_pp.freeze_panes = 'A2'

    # ── Sheet 3: Build Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Build Summary')
    hdr2 = ['Build', 'Executions', 'Passed', 'Failed', 'Pass %', 'First Run', 'Last Run']
    for ci, h in enumerate(hdr2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = HDR; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, row in enumerate(build_summary.itertuples(index=False), 2):
        vals = list(row)
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if ci == 5:  # Pass %
                ratio = float(val)
                cell.fill = GREEN if ratio >= 90 else ORANGE if ratio >= 70 else RED
                cell.value = f'{ratio:.1f}%'

    for ci, w in enumerate([10, 14, 12, 12, 12, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'A2'

    # ── Sheet 4: Summary stats ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Summary')
    summary_rows = [
        ('Version',            version),
        ('Min Build',          min_build),
        ('Builds Included',    ', '.join(str(b) for b in sorted(build_df['build'].tolist()))),
        ('Total Executions (Raw)',   int(build_df['executions'].sum())),
        ('Total Passed (Last Result)',  int((latest_df['status'] == 'passed').sum())),
        ('Total Failed (Last Result)',  int(latest_df['status'].isin(['failed', 'error', 'fail']).sum())),
        ('Overall Pass Rate (Last Result)', f'{int((latest_df["status"] == "passed").sum()) / max(len(latest_df), 1) * 100:.1f}%'),
        ('Critical Failures',  len(critical_df)),
        ('Generated',          datetime.now().strftime('%d %b %Y %H:%M')),
    ]
    for ri, (k, v) in enumerate(summary_rows, 1):
        ws3.cell(ri, 1, k).font  = bold_font
        ws3.cell(ri, 1, k).fill  = BLUE
        ws3.cell(ri, 2, v)
        for ci in (1, 2):
            ws3.cell(ri, ci).border    = thin_border
            ws3.cell(ri, ci).alignment = Alignment(horizontal='left')
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 55

    wb.save(xlsx_file)
    print(f'✓ Excel saved to {xlsx_file}')
    return xlsx_file


def main():
    parser = argparse.ArgumentParser(description='Automation build report')
    parser.add_argument('--version',   default=VERSION,   help='DP version (default: %(default)s)')
    parser.add_argument('--min-build', default=MIN_BUILD, type=int, help='Minimum build number (default: %(default)s)')
    args = parser.parse_args()

    version   = args.version
    min_build = args.min_build
    out_file  = f'automation_report_{version.replace(".", "_")}_build{min_build}plus.html'

    print(f'Automation Build Report — {version} builds ≥ {min_build}')
    print('-' * 55)

    print('Connecting to PostgreSQL...')
    conn = connect()
    print('✓ Connected\n')

    print('Fetching build summary...')
    build_df = get_build_summary(conn, version, min_build)
    print(f'✓ {len(build_df)} builds found: {sorted(build_df["build"].tolist())}')

    print('Fetching latest status per test × platform type...')
    latest_df = get_latest_status_by_platform_type(conn, version, min_build)
    print(f'✓ {len(latest_df)} latest execution records')

    prior_versions = get_prior_versions(version)
    print(f'Prior versions (baseline): {prior_versions}')

    print('Fetching baseline test counts...')
    baseline_df = get_available_tests_baseline(conn, prior_versions)

    print('Fetching new tests (not in prior releases)...')
    new_tests_df = get_new_tests(conn, version, min_build, prior_versions)
    print(f'✓ {new_tests_df["test_id"].nunique() if not new_tests_df.empty else 0} new tests found')

    print('Identifying critical failures...')
    critical_df = get_critical_failures(conn, version, min_build)
    print(f'✓ {len(critical_df)} critical failures')

    conn.close()

    print('\nGenerating HTML report...')
    html_content = build_html(version, min_build, build_df, latest_df, baseline_df, critical_df, new_tests_df, prior_versions)
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f'✓ Report saved to {out_file}')

    print('\nExporting critical failures to Excel...')
    xlsx_file = export_critical_failures_excel(critical_df, latest_df, build_df, baseline_df, version, min_build)

    # Open in default browser
    import webbrowser
    webbrowser.open(os.path.abspath(out_file))
    import subprocess
    subprocess.Popen(['start', '', os.path.abspath(xlsx_file)], shell=True)


if __name__ == '__main__':
    main()
